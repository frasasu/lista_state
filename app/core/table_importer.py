# table_importer.py
import csv
import openpyxl
import os
from typing import Dict, List, Optional, Tuple, Any

class TableImporter:
    """
    Importateur de fichiers CSV et Excel sans pandas.
    À intégrer directement dans votre classe ManagerApp.
    """
    
    def __init__(self, webview_instance):
        self.webview = webview_instance
    
    def import_table(self) -> Tuple[Optional[str], Optional[Dict[str, List]]]:
        """Importe un fichier CSV ou Excel via boîte de dialogue"""
        window = self.webview.active_window()
        file_types = ('Fichier csv(*.csv)', 'Fichier xlsx(*.xlsx)', 'Fichier xls(*.xls)')
        
        file_paths = window.create_file_dialog(
            self.webview.OPEN_DIALOG,
            allow_multiple=False,
            file_types=file_types
        )

        if not file_paths:
            return None, None

        file_path = file_paths[0]
        
        try:
            data = self._import_file(file_path)
            if data is None:
                return None, None
            
            filename = os.path.basename(file_path)
            name_without_ext = os.path.splitext(filename)[0]
            return name_without_ext, data
            
        except Exception as e:
            print(f"Erreur lors de l'import: {e}")
            return None, None
    
    def _import_file(self, file_path: str) -> Optional[Dict[str, List]]:
        """Importe un fichier depuis son chemin"""
        if file_path.endswith('.csv'):
            return self._import_csv(file_path)
        elif file_path.endswith(('.xlsx', '.xls')):
            return self._import_excel(file_path)
        return None
    
    def _import_csv(self, file_path: str) -> Dict[str, List]:
        """Importe un fichier CSV"""
        result = {}
        
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            lines = [line.strip() for line in f if line.strip()]
        
        if not lines:
            return result
        
        # Détecter le délimiteur
        first_line = lines[0]
        delimiter = self._detect_delimiter(first_line)
        
        # Lire les en-têtes
        headers = [h.strip() for h in first_line.split(delimiter)]
        
        # Initialiser les listes
        for header in headers:
            result[header] = []
        
        # Lire les données
        for line in lines[1:]:
            values = line.split(delimiter)
            for i, value in enumerate(values):
                if i < len(headers):
                    result[headers[i]].append(self._convert_value(value.strip()))
        
        return result
    
    def _import_excel(self, file_path: str) -> Dict[str, List]:
        """Importe un fichier Excel"""
        result = {}
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Lire les en-têtes
        headers = []
        for cell in ws[1]:
            header = cell.value or f"Colonne_{len(headers)+1}"
            headers.append(str(header))
        
        # Initialiser les listes
        for header in headers:
            result[header] = []
        
        # Lire les données
        for row in ws.iter_rows(min_row=2, values_only=True):
            for i, value in enumerate(row):
                if i < len(headers):
                    result[headers[i]].append(self._convert_value(value))
        
        return result
    
    def _detect_delimiter(self, line: str) -> str:
        """Détecte le délimiteur CSV"""
        delimiters = [',', ';', '\t', '|']
        counts = {d: line.count(d) for d in delimiters}
        return max(counts, key=counts.get) if max(counts.values()) > 0 else ','
    
    def _convert_value(self, value: Any) -> Any:
        """Convertit une valeur au bon type"""
        if value is None:
            return None
        if isinstance(value, (bool, int, float)):
            return value
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            if value.lower() in ('true', 'vrai', 'yes', 'oui', '1'):
                return True
            if value.lower() in ('false', 'faux', 'no', 'non', '0'):
                return False
            try:
                if value.isdigit() or (value[0] == '-' and value[1:].isdigit()):
                    return int(value)
            except:
                pass
            try:
                if '.' in value or 'e' in value.lower():
                    return float(value)
            except:
                pass
        return value